import io
import os
import textwrap

import openpyxl
from django.conf import settings
from django.utils import timezone

from ..constants import APP_TYPE_MAP
from ..utils import format_channels

# PDF出力時の余白設定 (cm単位)
PDF_MARGIN_TOP_CM = 1.0
PDF_MARGIN_BOTTOM_CM = 1.0
PDF_MARGIN_LEFT_CM = 0.6
PDF_MARGIN_RIGHT_CM = 0.6
PDF_MARGIN_HEADER_CM = 0.5
PDF_MARGIN_FOOTER_CM = 0.5

# PDF出力時の印刷範囲
PDF_PRINT_AREA = 'A1:AH81'

# 単位変換用定数 (1cm = 0.393701インチ)
CM_TO_INCH = 0.393701


class Cells:
    """Excel セル座標定数"""
    SUBMISSION_DATE = 'AD1'
    APP_TYPE = 'D4'
    MEMBER_ID_1 = 'M4'
    MEMBER_ID_2 = 'P4'
    MEMBER_NAME = 'W4'
    MEMBER_DEPT = 'M6'
    MEMBER_MANAGER = 'W6'
    MEMBER_PHONE = 'M8'
    MEMBER_EMAIL = 'W8'

    USER_NAME = 'O13'
    USER_TEL = 'L15'
    USER_EMAIL = 'W15'

    EVENT_NAME = 'H18'
    EVENT_COMMENT_LINES = ['E20', 'B21', 'B22']

    EXTRA_53CH = 'AF31'
    LMH_12G = 'AF27'

    # マイク数セル (Analog / Digital)
    MIC_COUNTS = {
        'analog_rm_10mw': 'K27',
        'analog_53ch_rm_10mw': 'R27',
        'analog_em_10mw': 'K28',
        'analog_53ch_em_10mw': 'R28',
        'digital_rm_10mw': 'K29',
        'digital_rm_20mw': 'M29',
        'digital_rm_50mw': 'O29',
        'digital_53ch_10mw': 'R29',
        'digital_53ch_20mw': 'U29',
        'digital_53ch_50mw': 'W29',
        'digital_12g_10mw': 'Z29',
        'digital_12g_20mw': 'AB29',
        'digital_12g_50mw': 'AD29',
    }


def _write_common_info(ws, data, member, current_date, pad_func):
    """共通情報（申請者、催事、マイク数など）をシートに書き込む"""

    # 0. 提出日
    ws[Cells.SUBMISSION_DATE] = current_date

    # 1. 基本情報
    app_type = data.get('app_type', 'new')
    ws[Cells.APP_TYPE] = APP_TYPE_MAP.get(app_type, '新規')

    if member:
        ws[Cells.MEMBER_ID_1] = member.member_id_1 or ''
        ws[Cells.MEMBER_ID_2] = member.member_id_2 or ''
        ws[Cells.MEMBER_NAME] = pad_func(member.name)
        ws[Cells.MEMBER_DEPT] = pad_func(member.department)
        ws[Cells.MEMBER_MANAGER] = pad_func(member.manager_name)
        ws[Cells.MEMBER_PHONE] = pad_func(member.phone)
        ws[Cells.MEMBER_EMAIL] = pad_func(member.email)

    # 2. 現地使用者
    user = data.get('user') or {}
    user_display = f'{user.get("name", "") or "未設定"}（{user.get("kana", "") or ""}）'
    ws[Cells.USER_NAME] = pad_func(user_display)
    ws[Cells.USER_TEL] = pad_func(user.get('tel', ''))
    ws[Cells.USER_EMAIL] = pad_func(user.get('email', ''))

    # 3. 催事情報
    event = data.get('event') or {}
    ws[Cells.EVENT_NAME] = pad_func(event.get('name', ''))

    # コメントの折り返し処理 (textwrapを使用して堅牢化)
    comment_raw = (event.get('comment') or '').replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
    comment_lines = textwrap.wrap(comment_raw, width=55)
    for i, cell_addr in enumerate(Cells.EVENT_COMMENT_LINES):
        val = comment_lines[i] if i < len(comment_lines) else ''
        ws[cell_addr] = pad_func(val)

    # 4. 使用マイク数
    mc = data.get('mic_counts') or {}
    ws[Cells.MIC_COUNTS['analog_rm_10mw']] = (mc.get('analog_rm') or {}).get('10mw', '')
    ws[Cells.MIC_COUNTS['analog_53ch_rm_10mw']] = (mc.get('analog_53ch') or {}).get('rm_10mw', '')
    ws[Cells.MIC_COUNTS['analog_em_10mw']] = (mc.get('analog_em') or {}).get('10mw', '')
    ws[Cells.MIC_COUNTS['analog_53ch_em_10mw']] = (mc.get('analog_53ch') or {}).get('em_10mw', '')

    ws[Cells.MIC_COUNTS['digital_rm_10mw']] = (mc.get('digital_rm') or {}).get('10mw', '')
    ws[Cells.MIC_COUNTS['digital_rm_20mw']] = (mc.get('digital_rm') or {}).get('20mw', '')
    ws[Cells.MIC_COUNTS['digital_rm_50mw']] = (mc.get('digital_rm') or {}).get('50mw', '')

    ws[Cells.MIC_COUNTS['digital_53ch_10mw']] = (mc.get('digital_53ch') or {}).get('10mw', '')
    ws[Cells.MIC_COUNTS['digital_53ch_20mw']] = (mc.get('digital_53ch') or {}).get('20mw', '')
    ws[Cells.MIC_COUNTS['digital_53ch_50mw']] = (mc.get('digital_53ch') or {}).get('50mw', '')

    ws[Cells.MIC_COUNTS['digital_12g_10mw']] = (mc.get('digital_12g') or {}).get('10mw', '')
    ws[Cells.MIC_COUNTS['digital_12g_20mw']] = (mc.get('digital_12g') or {}).get('20mw', '')
    ws[Cells.MIC_COUNTS['digital_12g_50mw']] = (mc.get('digital_12g') or {}).get('50mw', '')

    ws[Cells.LMH_12G] = mc.get('12g_lmh', '')

    if data.get('extra_53ch') == '○':
        ws[Cells.EXTRA_53CH] = '○'


def _write_facilities(ws, facilities_slice, pad_func):
    """施設リスト（最大4施設）をシートに書き込む"""
    for i, f in enumerate(facilities_slice):
        base_row = 34 + (i * 12)
        ws.cell(row=base_row, column=5).value = f.get('start_date', '')
        ws.cell(row=base_row, column=14).value = f.get('end_date', '')
        ws.cell(row=base_row, column=22).value = f.get('start_time', '')
        ws.cell(row=base_row, column=27).value = f.get('end_time', '')
        ws.cell(row=base_row + 2, column=10).value = pad_func(f.get('postal_code', ''))

        address_display = f'{f.get("prefecture", "") or ""}{f.get("address", "") or ""}'
        ws.cell(row=base_row + 2, column=18).value = pad_func(address_display)

        ws.cell(row=base_row + 4, column=12).value = pad_func(f.get('category', ''))
        ws.cell(row=base_row + 4, column=18).value = pad_func(f.get('name', ''))
        ws.cell(row=base_row + 6, column=15).value = pad_func(f.get('applied_area', ''))

        channels = f.get('selectedChannels')
        if not isinstance(channels, list):
            channels = []
        ws.cell(row=base_row + 8, column=15).value = pad_func(format_channels(channels))


def generate_adjustment_excel(data, member=None, for_pdf=False):
    """
    Excelテンプレートにデータを書き込んで返す。
    """
    template_path = os.path.join(settings.BASE_DIR, 'Excel', 'master.xlsx')
    if not os.path.exists(template_path):
        raise FileNotFoundError(f'Template not found at {template_path}')

    # テンプレート読み込み (計算済み値を優先的に扱う)
    wb = openpyxl.load_workbook(template_path, data_only=True)

    # 施設データの型安全性とフィルタリング
    raw_facilities = data.get('facilities') or []
    facilities = [f for f in raw_facilities if isinstance(f, dict)]

    num_facilities = len(facilities)

    # 使用するシートを決定
    target_sheets = ['master_01']
    if num_facilities > 4:
        target_sheets.append('master_02')
    if num_facilities > 8:
        target_sheets.append('master_03')

    # 日本時間での日付取得
    current_date = timezone.localtime().strftime('%Y/%m/%d')

    def pad(val):
        """PDF変換時のレイアウト崩れ防止用パディング"""
        if val is None:
            return ''
        s = str(val)
        # すでにスペースがある場合や数値セルへの考慮を入れた安全なパディング
        if for_pdf and s and not s.startswith(' '):
            return f' {s}'
        return s

    for sheet_idx, sheet_name in enumerate(target_sheets):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Required sheet '{sheet_name}' not found in template.")

        ws = wb[sheet_name]

        # PDF出力用のページ設定 (余白設定)
        if for_pdf:
            ws.page_margins.top = PDF_MARGIN_TOP_CM * CM_TO_INCH
            ws.page_margins.bottom = PDF_MARGIN_BOTTOM_CM * CM_TO_INCH
            ws.page_margins.left = PDF_MARGIN_LEFT_CM * CM_TO_INCH
            ws.page_margins.right = PDF_MARGIN_RIGHT_CM * CM_TO_INCH
            ws.page_margins.header = PDF_MARGIN_HEADER_CM * CM_TO_INCH
            ws.page_margins.footer = PDF_MARGIN_FOOTER_CM * CM_TO_INCH
            ws.print_area = PDF_PRINT_AREA

        # 共通情報の書き込み
        _write_common_info(ws, data, member, current_date, pad)

        # 施設情報の書き込み (シート毎に4つずつ)
        start_idx = sheet_idx * 4
        sheet_facilities = facilities[start_idx : start_idx + 4]
        _write_facilities(ws, sheet_facilities, pad)

    # 不要なシートを安全に削除
    for s in list(wb.sheetnames):
        if s not in target_sheets:
            del wb[s]

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
