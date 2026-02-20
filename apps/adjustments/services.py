import io
import openpyxl
import os
import subprocess
import tempfile
from datetime import datetime
from django.conf import settings

def format_channels(channels):
    if not channels: return ""
    sorted_channels = sorted(list(set(map(int, channels))))
    result = []
    i = 0
    while i < len(sorted_channels):
        start = sorted_channels[i]
        end = start
        while i + 1 < len(sorted_channels) and sorted_channels[i+1] == end + 1:
            end = sorted_channels[i+1]
            i += 1
        if end - start >= 2: result.append(f"{start}-{end}")
        else:
            for val in range(start, end + 1): result.append(str(val))
        i += 1
    return ", ".join(result)

def generate_adjustment_excel(data, member=None, for_pdf=False):
    """
    Excelテンプレートにデータを書き込んで返す。
    複数シート (master_01, master_02, master_03) に対応し、施設数に応じてシートを増やす。
    for_pdf=True の場合、PDF変換時のレイアウト調整として左寄せ項目に半角スペースを付与する。
    """
    template_path = os.path.join(settings.BASE_DIR, 'Excel', 'master.xlsx')
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found at {template_path}")
        
    wb = openpyxl.load_workbook(template_path)
    
    facilities = data.get('facilities', [])
    num_facilities = len(facilities)
    
    # 使用するシートを決定
    target_sheets = ['master_01']
    if num_facilities > 4:
        target_sheets.append('master_02')
    if num_facilities > 8:
        target_sheets.append('master_03')

    # 各シートに共通情報と施設情報を書き込む
    current_date = datetime.now().strftime('%Y/%m/%d')
    def pad(val):
        if not for_pdf: return val if val else ""
        return f" {val}" if val else ""

    for sheet_idx, sheet_name in enumerate(target_sheets):
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        
        # 0. 提出日 (AD1) - サーバーのロケールに依存せず YYYY/MM/DD 固定にするため上書き
        ws['AD1'] = current_date
        
        # 1. 基本情報 (共通)
        app_type = data.get('app_type', 'new')
        type_text = '新規' if app_type == 'new' else '変更' if app_type == 'change' else '削除'
        ws['D4'] = type_text 

        if member:
            ws['M4'] = member.member_id_1
            ws['P4'] = member.member_id_2
            ws['W4'] = pad(member.name)
            ws['M6'] = pad(member.department)
            ws['W6'] = pad(member.manager_name)
            ws['M8'] = pad(member.phone)
            ws['W8'] = pad(member.email)

        # 2. 現地使用者 (共通)
        user = data.get('user', {})
        ws['O13'] = pad(f"{user.get('name', '')}（{user.get('kana', '')}）")
        ws['L15'] = pad(user.get('tel', ''))
        ws['W15'] = pad(user.get('email', ''))

        # 3. 催事情報 (共通)
        event = data.get('event', {})
        ws['H18'] = pad(event.get('name', ''))
        
        comment = (event.get('comment') or "").replace('\n', '').replace('\r', '')
        ws['E20'] = pad(comment[0:55])
        ws['B21'] = pad(comment[55:110])
        ws['B22'] = pad(comment[110:165])

        # 4. 使用マイク数 (共通)
        mc = data.get('mic_counts', {})
        ws['K27'] = mc.get('analog_rm', {}).get('10mw', '')
        ws['R27'] = mc.get('analog_53ch', {}).get('rm_10mw', '')
        ws['K28'] = mc.get('analog_em', {}).get('10mw', '')
        ws['R28'] = mc.get('analog_53ch', {}).get('em_10mw', '')
        
        ws['K29'] = mc.get('digital_rm', {}).get('10mw', '')
        ws['M29'] = mc.get('digital_rm', {}).get('20mw', '')
        ws['O29'] = mc.get('digital_rm', {}).get('50mw', '')
        ws['R29'] = mc.get('digital_53ch', {}).get('10mw', '')
        ws['U29'] = mc.get('digital_53ch', {}).get('20mw', '')
        ws['W29'] = mc.get('digital_53ch', {}).get('50mw', '')
        
        ws['Z29'] = mc.get('digital_12g', {}).get('10mw', '')
        ws['AB29'] = mc.get('digital_12g', {}).get('20mw', '')
        ws['AD29'] = mc.get('digital_12g', {}).get('50mw', '')
        ws['AF27'] = mc.get('12g_lmh', '')

        if data.get('extra_53ch') == '○':
            ws['AF31'] = '○'

        # 5. 施設リスト (シート毎に4つずつ)
        start_idx = sheet_idx * 4
        sheet_facilities = facilities[start_idx : start_idx + 4]
        
        for i, f in enumerate(sheet_facilities):
            base_row = 34 + (i * 12)
            ws.cell(row=base_row, column=5).value = f.get('start_date', '')
            ws.cell(row=base_row, column=14).value = f.get('end_date', '')
            ws.cell(row=base_row, column=22).value = f.get('start_time', '')
            ws.cell(row=base_row, column=27).value = f.get('end_time', '')
            ws.cell(row=base_row + 2, column=10).value = pad(f.get('postal_code', ''))
            ws.cell(row=base_row + 2, column=18).value = pad(f"{f.get('prefecture', '')}{f.get('address', '')}")
            ws.cell(row=base_row + 4, column=12).value = pad(f.get('category', ''))
            ws.cell(row=base_row + 4, column=18).value = pad(f.get('name', ''))
            ws.cell(row=base_row + 6, column=15).value = pad(f.get('applied_area', ''))
            ws.cell(row=base_row + 8, column=15).value = pad(format_channels(f.get('selectedChannels', [])))

    # 不要なシートを削除 (PDF出力時に含まれないようにする)
    all_sheets = wb.sheetnames
    for s in all_sheets:
        if s not in target_sheets:
            del wb[s]

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def convert_excel_to_pdf(excel_buffer):
    """
    LibreOfficeを使用してExcelバッファをPDFバッファに変換する。
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        excel_path = os.path.join(tmp_dir, "temp.xlsx")
        with open(excel_path, "wb") as f:
            f.write(excel_buffer.read())
        
        # LibreOfficeによる変換 (headlessモード)
        # soffice または libreoffice コマンドを試す
        commands = ['libreoffice', 'soffice']
        success = False
        last_error = ""

        for cmd in commands:
            try:
                result = subprocess.run([
                    cmd, '--headless', '--convert-to', 'pdf',
                    '--outdir', tmp_dir, excel_path
                ], check=True, capture_output=True, text=True)
                success = True
                break
            except (subprocess.CalledProcessError, FileNotFoundError) as e:
                last_error = str(e)
                if isinstance(e, subprocess.CalledProcessError):
                    last_error += f"\nStderr: {e.stderr}"
                continue
        
        if not success:
            raise RuntimeError(f"PDF conversion failed. Ensure LibreOffice is installed. Error: {last_error}")
            
        pdf_path = os.path.join(tmp_dir, "temp.pdf")
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"Conversion failed: PDF file not created at {pdf_path}")
            
        with open(pdf_path, "rb") as f:
            pdf_buffer = io.BytesIO(f.read())
        
        pdf_buffer.seek(0)
        return pdf_buffer

def generate_adjustment_pdf(data, member=None):
    """
    データをExcelに転記し、それをPDFに変換する (Single Source of Truth 方式)。
    """
    excel_buffer = generate_adjustment_excel(data, member, for_pdf=True)
    return convert_excel_to_pdf(excel_buffer)
