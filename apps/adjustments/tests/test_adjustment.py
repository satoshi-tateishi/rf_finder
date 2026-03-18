import io
import json
import subprocess
import unittest.mock
from unittest.mock import MagicMock, patch

from django.core import mail
from django.test import TestCase
from django.urls import reverse

from apps.accounts.models import EmailTemplate, Member
from apps.adjustments.services import (
    generate_adjustment_excel,
    generate_adjustment_pdf,
    send_adjustment_email,
)
from apps.adjustments.services.pdf_service import convert_excel_to_pdf
from apps.adjustments.utils import format_channels


class AdjustmentLogicTest(TestCase):
    def test_format_channels(self):
        """チャンネル番号リストが正しいハイフン連結文字列になること"""
        self.assertEqual(format_channels([13, 14, 15]), '13-15')
        self.assertEqual(format_channels([13, 14, 16, 17, 18]), '13, 14, 16-18')
        self.assertEqual(format_channels([13, 15, 17]), '13, 15, 17')
        self.assertEqual(format_channels([]), '')

    @patch('apps.adjustments.services.pdf_service.convert_excel_to_pdf')
    def test_generate_excel_and_pdf_smoke(self, mock_convert):
        """ExcelとPDFの生成がエラーなく完了し、BytesIOを返すこと"""
        mock_convert.return_value = io.BytesIO(b'dummy pdf content')
        member = Member.objects.create(
            member_id_1='123',
            member_id_2='4567',
            name='テスト会員',
            manager_name='担当者',
            phone='03-1234-5678',
            email='test@example.com',
        )
        data = {
            'app_type': 'new',
            'user': {'name': '使用者', 'kana': 'しようしゃ', 'tel': '090', 'email': 'u@ex.com'},
            'event': {'name': '催事', 'comment': 'コメント'},
            'facilities': [
                {'name': '施設1', 'start_date': '2026-02-20', 'start_time': '09:00', 'selectedChannels': [13, 14]}
            ],
            'mic_counts': {'analog_rm': {'10mw': 1}},
        }

        # Excel生成テスト
        excel_buffer = generate_adjustment_excel(data, member)
        self.assertIsInstance(excel_buffer, io.BytesIO)
        self.assertTrue(len(excel_buffer.getvalue()) > 0)

        # PDF生成テスト
        try:
            pdf_buffer = generate_adjustment_pdf(data, member)
            self.assertIsInstance(pdf_buffer, io.BytesIO)
            self.assertEqual(pdf_buffer.getvalue(), b'dummy pdf content')
        except Exception as e:
            self.fail(f'PDF generation failed: {e}')

    def test_send_adjustment_email(self):
        """メール送信がテンプレートに基づいて正しく行われること"""
        member = Member.objects.create(name='テスト会社')
        EmailTemplate.objects.create(
            subject='【{運用日}】テスト',
            body='こんにちは {ユーザー名} 様。区分は {タイプ} です。',
            to_address='recipient@example.com',
            cc_address='{ユーザーEメールアドレス}, boss@example.com',
        )
        data = {
            'app_type': 'change',
            'user': {'name': '太郎', 'email': 'taro@ex.com'},
            'facilities': [{'start_date': '2026-02-20'}],
        }
        pdf_buffer = io.BytesIO(b'dummy pdf content')

        send_adjustment_email(data, member, pdf_buffer)

        # 送信済みメールの検証
        self.assertEqual(len(mail.outbox), 1)
        sent = mail.outbox[0]
        self.assertEqual(sent.subject, '【2026年2月20日】テスト')
        self.assertIn('こんにちは 太郎 様。区分は 変更 です。', sent.body)
        # CCの検証
        self.assertIn('taro@ex.com', sent.cc)
        self.assertIn('boss@example.com', sent.cc)
        self.assertEqual(len(sent.attachments), 1)
        # MIMEApplication の場合は get_filename() などで取得
        attachment = sent.attachments[0]
        filename = attachment.get_filename()
        self.assertIn('運用連絡票', filename)
        self.assertIn('20260220', filename)


class AdjustmentAPITest(TestCase):
    def setUp(self):
        # PDF変換のモック化
        self.pdf_patcher = patch('apps.adjustments.services.pdf_service.convert_excel_to_pdf')
        self.mock_convert = self.pdf_patcher.start()
        self.mock_convert.return_value = io.BytesIO(b'dummy pdf')

        # LINE Bot サービスのモック化（実APIを呼ばないようにする）
        self.line_bot_patcher = patch('apps.adjustments.views.LineBotService')
        self.mock_line_bot = self.line_bot_patcher.start()

        from django.contrib.auth.models import User

        from apps.accounts.models import UserProfile

        # テスト用ユーザーの作成とログイン
        self.user = User.objects.create_user(username='testuser', password='password')
        # UserProfile はシグナルで自動生成されるが、role などを明示的に設定
        self.user.profile.role = UserProfile.Role.EDITOR
        self.user.profile.family_name = 'テスト'
        self.user.profile.given_name = 'ユーザー'
        self.user.profile.save()
        self.client.force_login(self.user)

        Member.objects.create(name='テスト会員')
        EmailTemplate.objects.create(subject='テスト件名', body='テスト本文', to_address='test@example.com')
        self.valid_data = {
            'app_type': 'new',
            'user': {'name': '使用者', 'kana': 'しようしゃ', 'tel': '090-1234-5678', 'email': 'u@ex.com'},
            'event': {'name': '催事'},
            'facilities': [
                {
                    'name': '施設1',
                    'start_date': '2026-02-20',
                    'end_date': '2026-02-20',
                    'start_time': '09:00',
                    'end_time': '22:00',
                }
            ],
            'mic_counts': {'analog_rm': {'10mw': 1}},
        }

    def tearDown(self):
        self.pdf_patcher.stop()
        self.line_bot_patcher.stop()

    def test_preview_pdf_api(self):
        """PDFプレビューAPIが正常にPDFを返すこと"""
        response = self.client.post(
            reverse('adjustments:preview_pdf'), data=json.dumps(self.valid_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_preview_excel_api(self):
        """ExcelプレビューAPIが正常にExcelファイルを返すこと"""
        response = self.client.post(
            reverse('adjustments:preview_excel'), data=json.dumps(self.valid_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_send_email_api(self):
        """メール送信APIが正常に受理されること"""
        response = self.client.post(
            reverse('adjustments:send_email'), data=json.dumps(self.valid_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'success')
        self.assertEqual(len(mail.outbox), 1)

    def test_save_adjustment_api(self):
        """下書き保存APIが正常に動作し、DBに保存されること"""
        from apps.adjustments.models import OperationAdjustment

        response = self.client.post(
            reverse('adjustments:save_adjustment'), data=json.dumps(self.valid_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'success')
        adj_id = response.json()['data']['id']

        # DB確認
        adj = OperationAdjustment.objects.get(pk=adj_id)
        self.assertEqual(adj.event_name, '催事')
        self.assertEqual(adj.status, 'draft')
        self.assertEqual(adj.user_name, '使用者')

    def test_api_invalid_json(self):
        """不正なJSONを送った場合に400エラーになること"""
        response = self.client.post(
            reverse('adjustments:preview_pdf'), data='invalid json', content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['status'], 'error')

    def test_missing_member(self):
        """会員情報（Member）がDBに存在しない場合でも生成処理が継続されること"""
        Member.objects.all().delete()
        response = self.client.post(
            reverse('adjustments:preview_pdf'), data=json.dumps(self.valid_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_validation_error_empty_payload(self):
        """空のJSONオブジェクトを送った場合にバリデーションエラーになること"""
        response = self.client.post(
            reverse('adjustments:preview_pdf'), data=json.dumps({}), content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['status'], 'error')
        self.assertIn('errors', response.json())

    def test_validation_error_missing_event_name(self):
        """必須項目（催事名）が欠落している場合にエラーになること"""
        incomplete_data = self.valid_data.copy()
        incomplete_data['event'] = {}  # 催事名がない

        response = self.client.post(
            reverse('adjustments:preview_pdf'), data=json.dumps(incomplete_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['status'], 'error')
        # エラーキーが event_name になっていることを確認 (UserInfoForm/EventInfoFormの接頭辞)
        self.assertIn('event_name', response.json()['errors'])

    def test_validation_error_missing_user_fields(self):
        """現地使用者の必須項目が欠落している場合にエラーになること"""
        incomplete_data = self.valid_data.copy()
        incomplete_data['user'] = {'name': '名前のみ'}  # 他が足りない

        response = self.client.post(
            reverse('adjustments:preview_pdf'), data=json.dumps(incomplete_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        errors = response.json()['errors']
        self.assertIn('user_kana', errors)
        self.assertIn('user_tel', errors)
        self.assertIn('user_email', errors)

    def test_validation_error_no_mic_counts(self):
        """マイク数が1つも入力されていない場合にエラーになること"""
        incomplete_data = self.valid_data.copy()
        incomplete_data['mic_counts'] = {}  # 空

        response = self.client.post(
            reverse('adjustments:preview_pdf'), data=json.dumps(incomplete_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('mic_counts', response.json()['errors'])

    def test_validation_error_incomplete_facility(self):
        """施設の時間情報が不足している場合にエラーになること"""
        incomplete_data = self.valid_data.copy()
        incomplete_data['facilities'] = [{'name': '施設1', 'start_date': '2026-02-20'}]  # 時間がない

        response = self.client.post(
            reverse('adjustments:preview_pdf'), data=json.dumps(incomplete_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('facilities', response.json()['errors'])

    def test_validation_error_zero_mics(self):
        """マイク数がすべて0の場合にエラーになること"""
        invalid_data = self.valid_data.copy()
        invalid_data['mic_counts'] = {'analog_rm': {'10mw': 0}, 'digital_rm': {'10mw': '0', '20mw': 0}}
        response = self.client.post(
            reverse('adjustments:preview_pdf'), data=json.dumps(invalid_data), content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('mic_counts', response.json()['errors'])


class AdjustmentUtilityTest(TestCase):
    def test_get_adjustment_filename_sanitization(self):
        """ファイル名に使用できない文字がサニタイズされること"""
        from apps.adjustments.utils import get_adjustment_filename

        data = {
            'app_type': 'change',
            'event': {'name': '催事 / ハムレット : "復讐" *'},
            'facilities': [{'start_date': '2026-02-22'}],
        }
        filename = get_adjustment_filename(data, 'pdf')
        # 禁止文字 / : " * が除外され、ファイル名が妥当であることを確認
        self.assertNotIn('/', filename)
        self.assertNotIn(':', filename)
        self.assertNotIn('"', filename)
        self.assertNotIn('*', filename)
        self.assertIn('運用連絡票_変更_', filename)
        self.assertIn('20260222.pdf', filename)

    def test_get_adjustment_filename_app_types(self):
        """各申請区分が正しく日本語に変換されること"""
        from apps.adjustments.utils import get_adjustment_filename

        base_data = {'event': {'name': 'テスト'}, 'facilities': []}

        self.assertIn('新規', get_adjustment_filename({**base_data, 'app_type': 'new'}))
        self.assertIn('変更', get_adjustment_filename({**base_data, 'app_type': 'change'}))
        self.assertIn('削除', get_adjustment_filename({**base_data, 'app_type': 'delete'}))


class EmailServiceDetailTest(TestCase):
    def test_template_full_replacement(self):
        """すべてのプレースホルダーが正しく置換されること"""
        from apps.adjustments.services import send_adjustment_email

        member = Member.objects.create(name='テスト会員')
        EmailTemplate.objects.create(
            subject='{タイプ}:{催事名}',
            body='{ユーザー名} 様 ({ユーザーEメールアドレス})\n運用日: {運用日}',
            to_address='to@ex.com',
        )
        data = {
            'app_type': 'new',
            'user': {'name': '太郎', 'email': 'taro@ex.com'},
            'event': {'name': 'ハムレット'},
            'facilities': [{'start_date': '2026-05-10'}],
        }
        pdf_buffer = io.BytesIO(b'dummy')

        send_adjustment_email(data, member, pdf_buffer)

        sent = mail.outbox[0]
        self.assertEqual(sent.subject, '新規:ハムレット')
        self.assertIn('太郎 様 (taro@ex.com)', sent.body)
        self.assertIn('運用日: 2026年5月10日', sent.body)


class ConvertExcelToPdfTest(TestCase):
    """convert_excel_to_pdf のユニットテスト。

    subprocess.run を mock してコマンド実行を伴わずに
    各シナリオ（成功・フォールバック・失敗・ファイル未生成）を検証する。
    """

    def _make_excel_buffer(self):
        """ダミーの xlsx バッファを返す"""
        buf = io.BytesIO(b'dummy xlsx content')
        buf.seek(0)
        return buf

    @patch('apps.adjustments.services.pdf_service.subprocess.run')
    @patch('apps.adjustments.services.pdf_service.os.path.exists', return_value=True)
    def test_successful_conversion(self, mock_exists, mock_run):
        """subprocess が成功し PDF ファイルが生成されたとき BytesIO を返す"""
        mock_run.return_value = MagicMock(returncode=0)

        with patch('builtins.open', unittest.mock.mock_open(read_data=b'%PDF-dummy')):
            result = convert_excel_to_pdf(self._make_excel_buffer())

        self.assertIsInstance(result, io.BytesIO)

    @patch('apps.adjustments.services.pdf_service.subprocess.run')
    @patch('apps.adjustments.services.pdf_service.os.path.exists', return_value=True)
    def test_fallback_to_soffice(self, mock_exists, mock_run):
        """`libreoffice` が FileNotFoundError → `soffice` で成功するフォールバック"""
        def side_effect(cmd, **kwargs):
            if cmd[0] == 'libreoffice':
                raise FileNotFoundError('libreoffice not found')
            return MagicMock(returncode=0)

        mock_run.side_effect = side_effect

        with patch('builtins.open', unittest.mock.mock_open(read_data=b'%PDF-dummy')):
            result = convert_excel_to_pdf(self._make_excel_buffer())

        self.assertIsInstance(result, io.BytesIO)
        called_cmds = [c[0][0][0] for c in mock_run.call_args_list]
        self.assertIn('soffice', called_cmds)

    @patch('apps.adjustments.services.pdf_service.subprocess.run')
    def test_all_commands_fail_raises_runtime_error(self, mock_run):
        """`libreoffice` も `soffice` も FileNotFoundError → RuntimeError"""
        mock_run.side_effect = FileNotFoundError('command not found')

        with patch('builtins.open', unittest.mock.mock_open()):
            with self.assertRaises(RuntimeError) as ctx:
                convert_excel_to_pdf(self._make_excel_buffer())

        self.assertIn('PDF conversion failed', str(ctx.exception))

    @patch('apps.adjustments.services.pdf_service.subprocess.run')
    def test_exit_code_77_raises_runtime_error(self, mock_run):
        """exit code 77（User installation could not be completed）→ RuntimeError

        Dockerfile の appuser ホームディレクトリ未作成時に発生した障害の再発防止。
        """
        mock_run.side_effect = subprocess.CalledProcessError(
            returncode=77,
            cmd=['soffice', '--headless', '--convert-to', 'pdf'],
            stderr='Fatal Error: The application cannot be started. '
                   'User installation could not be completed.',
        )

        with patch('builtins.open', unittest.mock.mock_open()):
            with self.assertRaises(RuntimeError) as ctx:
                convert_excel_to_pdf(self._make_excel_buffer())

        self.assertIn('PDF conversion failed', str(ctx.exception))

    @patch('apps.adjustments.services.pdf_service.subprocess.run')
    @patch('apps.adjustments.services.pdf_service.os.path.exists', return_value=False)
    def test_pdf_file_not_created_raises_file_not_found(self, mock_exists, mock_run):
        """subprocess 成功でも PDF ファイルが存在しない → FileNotFoundError"""
        mock_run.return_value = MagicMock(returncode=0)

        with patch('builtins.open', unittest.mock.mock_open()):
            with self.assertRaises(FileNotFoundError):
                convert_excel_to_pdf(self._make_excel_buffer())


class EmailServiceEdgeCaseTest(TestCase):
    """email_service のエッジケーステスト"""

    BASE_DATA = {
        'app_type': 'new',
        'user': {'name': '太郎', 'email': 'taro@ex.com'},
        'event': {'name': 'テスト催事'},
        'facilities': [{'start_date': '2026-05-10'}],
    }

    def test_no_template_raises_runtime_error(self):
        """EmailTemplate が1件も存在しない場合 RuntimeError が発生すること"""
        member = Member.objects.create(name='会員')
        pdf_buffer = io.BytesIO(b'dummy')

        with self.assertRaises(RuntimeError) as ctx:
            send_adjustment_email(self.BASE_DATA, member, pdf_buffer)

        self.assertIn('メールテンプレートが登録されていません', str(ctx.exception))

    def test_empty_to_address_raises_runtime_error(self):
        """to_address が空のテンプレートで RuntimeError が発生すること"""
        member = Member.objects.create(name='会員')
        EmailTemplate.objects.create(
            subject='テスト',
            body='本文',
            to_address='',  # 空
        )
        pdf_buffer = io.BytesIO(b'dummy')

        with self.assertRaises(RuntimeError) as ctx:
            send_adjustment_email(self.BASE_DATA, member, pdf_buffer)

        self.assertIn('送信先メールアドレス', str(ctx.exception))

    def test_invalid_to_address_raises_runtime_error(self):
        """不正形式の to_address で RuntimeError が発生すること"""
        member = Member.objects.create(name='会員')
        EmailTemplate.objects.create(
            subject='テスト',
            body='本文',
            to_address='not-an-email',  # 不正形式
        )
        pdf_buffer = io.BytesIO(b'dummy')

        with self.assertRaises(RuntimeError) as ctx:
            send_adjustment_email(self.BASE_DATA, member, pdf_buffer)

        self.assertIn('不正なメールアドレス形式', str(ctx.exception))

    def test_member_name_placeholder_with_none_member(self):
        """member=None のとき {会員名} と {運用担当者} が空文字に置換されること"""
        EmailTemplate.objects.create(
            subject='テスト',
            body='会員名:{会員名} 担当者:{運用担当者}',
            to_address='to@ex.com',
        )
        pdf_buffer = io.BytesIO(b'dummy')

        send_adjustment_email(self.BASE_DATA, None, pdf_buffer)

        sent = mail.outbox[0]
        # プレースホルダーが空文字に置換されて残っていないこと
        self.assertNotIn('{会員名}', sent.body)
        self.assertNotIn('{運用担当者}', sent.body)
        self.assertIn('会員名:', sent.body)
        self.assertIn('担当者:', sent.body)


class ExcelServiceCellTest(TestCase):
    """generate_adjustment_excel のセル値書き込みテスト"""

    def setUp(self):
        self.member = Member.objects.create(
            member_id_1='001',
            member_id_2='0001',
            name='テスト会員',
            department='音響部',
            manager_name='田中一郎',
            phone='03-0000-0000',
            email='member@ex.com',
        )
        self.data = {
            'app_type': 'new',
            'user': {'name': '現地使用者', 'kana': 'げんちしようしゃ', 'tel': '090-0000-0001', 'email': 'user@ex.com'},
            'event': {'name': 'テスト催事', 'comment': 'テストコメント'},
            'facilities': [
                {
                    'name': '施設A',
                    'prefecture': '東京都',
                    'address': '渋谷区1-1',
                    'postal_code': '150-0001',
                    'applied_area': 'A帯',
                    'category': '屋内',
                    'start_date': '2026-06-01',
                    'end_date': '2026-06-02',
                    'start_time': '10:00',
                    'end_time': '21:00',
                    'selectedChannels': [13, 14, 15],
                }
            ],
            'mic_counts': {'analog_rm': {'10mw': 3}},
        }

    def _load_sheet(self, buffer):
        import openpyxl
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer, data_only=True)
        return wb['master_01']

    def test_user_info_cells(self):
        """現地使用者の名前・電話・メールが正しいセルに書き込まれること"""
        buf = generate_adjustment_excel(self.data, self.member)
        ws = self._load_sheet(buf)

        self.assertIn('現地使用者', ws['O13'].value)
        self.assertIn('090-0000-0001', ws['L15'].value)
        self.assertIn('user@ex.com', ws['W15'].value)

    def test_event_name_cell(self):
        """催事名が正しいセルに書き込まれること"""
        buf = generate_adjustment_excel(self.data, self.member)
        ws = self._load_sheet(buf)

        self.assertIn('テスト催事', ws['H18'].value)

    def test_member_info_cells(self):
        """会員情報（ID・名前・担当者）が正しいセルに書き込まれること"""
        buf = generate_adjustment_excel(self.data, self.member)
        ws = self._load_sheet(buf)

        self.assertEqual(ws['M4'].value, '001')
        self.assertEqual(ws['P4'].value, '0001')
        self.assertIn('テスト会員', ws['W4'].value)
        self.assertIn('田中一郎', ws['W6'].value)

    def test_mic_count_cell(self):
        """マイク数（analog_rm 10mw = 3）が正しいセルに書き込まれること"""
        buf = generate_adjustment_excel(self.data, self.member)
        ws = self._load_sheet(buf)

        self.assertEqual(ws['K27'].value, 3)

    def test_app_type_new(self):
        """申請区分「新規」が APP_TYPE セルに書き込まれること"""
        buf = generate_adjustment_excel(self.data, self.member)
        ws = self._load_sheet(buf)

        self.assertEqual(ws['D4'].value, '新規')

    def test_template_not_found_raises_error(self):
        """テンプレートファイルが存在しない場合 FileNotFoundError が発生すること"""
        with patch('apps.adjustments.services.excel_service.os.path.exists', return_value=False):
            with self.assertRaises(FileNotFoundError) as ctx:
                generate_adjustment_excel(self.data, self.member)

        self.assertIn('Template not found', str(ctx.exception))

    def test_multi_facility_single_sheet(self):
        """施設数 ≤ 4 のとき master_01 シートのみ出力されること"""
        buf = generate_adjustment_excel(self.data, self.member)
        buf.seek(0)
        import openpyxl
        wb = openpyxl.load_workbook(buf, data_only=True)
        self.assertIn('master_01', wb.sheetnames)
        self.assertNotIn('master_02', wb.sheetnames)
